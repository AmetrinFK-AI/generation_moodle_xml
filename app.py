import re
import os
import subprocess
import tempfile
import glob
from io import BytesIO

import streamlit as st
import openai
from openpyxl import load_workbook
from docx import Document

# ================== Налаштування ==================
# Встановіть свій ключ OpenAI через змінну середовища (OPENAI_API_KEY)
openai.api_key = os.getenv("OPENAI_API_KEY")

# ================== Налаштування сторінки ==================
st.set_page_config(page_title="Генератор тестів для Moodle (XML)", layout="wide")
st.markdown("""
<style>
textarea, input[type=text] {
    border-radius: 8px;
    padding: 8px;
    font-size: 1rem;
}
div[data-testid="column"] > div {
    background-color: #f9f9f9;
    padding: 1rem;
    margin-bottom: 1rem;
    border-radius: 8px;
    box-shadow: 0 0 10px rgba(0,0,0,0.1);
}
</style>
""", unsafe_allow_html=True)
st.title("Генератор тестів для Moodle (XML)")

with st.expander("📖 Інструкція до роботи з програмою", expanded=False):
    st.markdown("""
**Excel-режим**
- Підготуйте Excel-файл (.xlsx).
- Колонка A: текст питання.
- Колонка A: варіанти відповідей.
- Позначте правильні відповіді жовтим фоном (FFFF00).

**GPT-режим**
- Вставте текст українською мовою.
- Система створить 10 питань:
  - 4–5 Single-choice,
  - 2–3 True/False,
  - 2–3 Multiple-choice.
- Для кожного, крім True/False, 4 варіанти A–D.

**Ручне створення**
- Оберіть тип питання: Single-choice, Multiple-choice, True/False.
- Введіть текст питання та варіанти.
- Позначте правильні відповіді.

**Готовий тест**
- Формат:

1. Питання...
A. Відповідь A  
B. Відповідь B  
C. Відповідь C  
D. Відповідь D  
Правильний відповідь: B

- Для True/False:

7. Питання...  
Варіанти: True / False  
Правильний відповідь: True

- Обов'язкові: нумерація, варіанти, рядок з `Правильний відповідь:`.

**Word-режим**
    - **Правильний** варіант виділяйте **жирним** лише текст відповіді (не префікс).
    - Якщо ваше питання займає кілька рядків, тримайте їх в одному параграфі до першої відповіді.
    - **Не** використовуйте inline-список через `:` і `;` — кожен варіант має власний параграф.

    Приклад:
    1. Який термін подачі пропозиції по процедурі запит (ціни) пропозиції?
    A. від 2 днів
    B. від 3 днів
    C. від 5 днів
    2. Скільки днів після оголошення закупівлі тривають відкриті торги з особливостями:
    A. 14 днів з дня оприлюднення в електронній системі закупівель
    B. за рішенням замовника
    C. не раніше ніж за сім днів
    3. Після спливу якого терміну для підписання договору замовник має право відхилити пропозицію учасника
    А. після спливу 5 днів (на 6 день)
    A. після спливу 6 днів
    С. . після спливу 3 днів (на 4 день)

**YouTube to XML**
    Просто вставте посилання на відео та чекайте результату
""", unsafe_allow_html=False)

# ================== Утиліти для XML ==================

def wrap_cdata(text: str) -> str:
    """Обгортає текст у CDATA з HTML-тегом <p>."""
    return f"<![CDATA[<p>{text}</p>]]>"

def detect_question_type(answers):
    """Визначаємо тип питання за списком відповідей."""
    if all('-' in ans for ans, _ in answers):
        return "matching"
    correct = sum(1 for _, c in answers if c)
    if len(answers) == 2 and correct <= 1:
        return "truefalse"
    if correct == 1:
        return "single"
    if correct > 1:
        return "multiple"
    return "unknown"

def generate_moodle_xml_string(questions) -> str:
    """Генерує Moodle XML зі списку питань."""
    lines = ['<?xml version="1.0" encoding="UTF-8"?>', '<quiz>']
    for q in questions:
        q_type = detect_question_type(q["answers"])
        if q_type in ("single", "multiple"):
            lines.append('  <question type="multichoice">')
        elif q_type == "truefalse":
            lines.append('  <question type="truefalse">')
        elif q_type == "matching":
            lines.append('  <question type="matching">')
        else:
            continue

        preview = q["text"][:30] + ('...' if len(q["text"]) > 30 else '')
        lines.extend([
            '    <name>',
            f'      <text>{wrap_cdata(preview)}</text>',
            '    </name>',
            '    <questiontext format="html">',
            f'      <text>{wrap_cdata(q["text"])}</text>',
            '    </questiontext>'
        ])

        if q_type in ("single", "multiple"):
            total_correct = sum(1 for _, c in q["answers"] if c)
            penalty = 1.0 / total_correct if total_correct else 0
            lines.extend([
                '    <shuffleanswers>true</shuffleanswers>',
                f'    <single>{"true" if q_type=="single" else "false"}</single>',
                '    <answernumbering>abc</answernumbering>',
                f'    <penalty>{penalty:.6f}</penalty>',
                '    <defaultgrade>1.000000</defaultgrade>'
            ])
            for text, corr in q["answers"]:
                frac = 100 if corr else 0
                lines.extend([
                    f'    <answer fraction="{frac}" format="html">',
                    f'      <text><![CDATA[{text}]]></text>',
                    '    </answer>'
                ])

        elif q_type == "truefalse":
            correct_true = q["answers"][0][1]
            for val in ("true", "false"):
                frac = 100 if (val=="true" and correct_true) or (val=="false" and not correct_true) else 0
                lines.extend([
                    f'    <answer fraction="{frac}" format="html">',
                    f'      <text><![CDATA[{val}]]></text>',
                    '    </answer>'
                ])

        elif q_type == "matching":
            lines.append('    <shuffleanswers>true</shuffleanswers>')
            for pair, _ in q["answers"]:
                left, right = map(str.strip, pair.split('-', 1))
                lines.extend([
                    '    <subquestion format="html">',
                    f'      <text><![CDATA[{left}]]></text>',
                    f'      <answer><![CDATA[{right}]]></answer>',
                    '    </subquestion>'
                ])

        lines.append('  </question>')
    lines.append('</quiz>')
    return "\n".join(lines)

def download_xml(data, filename: str):
    """Створює кнопку для завантаження XML-файлу."""
    xml_bytes = data.encode('utf-8') if isinstance(data, str) else data
    st.download_button(
        label="📥 Завантажити XML",
        data=xml_bytes,
        file_name=filename,
        mime="application/xml"
    )

# ================== Утиліта для YouTube ==================

def download_audio_from_youtube(url: str) -> str:
    """
    Завантажує оригінальний аудіо-потік з YouTube за допомогою yt-dlp
    без пост-обробки ffmpeg, і повертає шлях до файлу.
    """
    tmpdir = tempfile.mkdtemp()
    # шаблон: збережемо файл у тому вигляді, в якому він був у джерелі
    out_template = os.path.join(tmpdir, 'audio.%(ext)s')
    result = subprocess.run(
        ['yt-dlp', '-f', 'bestaudio', '-o', out_template, url],
        capture_output=True,
        text=True
    )
    if result.returncode != 0:
        raise Exception(f"Помилка завантаження аудіо: {result.stderr}")
    files = glob.glob(os.path.join(tmpdir, 'audio.*'))
    if not files:
        raise Exception("Не знайдено файлу аудіо після завантаження")
    return files[0]


# ================== Парсери ==================

def parse_text_format(text: str):
    """Парсер готового тесту з тексту."""
    mapping = {'а':'A','А':'A','б':'B','Б':'B','в':'C','В':'C','г':'D','Г':'D'}
    blocks = re.split(r"(?m)(?=^\d+\.)", text.strip())
    blocks = [blk for blk in blocks if blk.strip()]
    corr_pattern = re.compile(r"(?i)^(?:правильн\w*\s+(?:ответ|відповід))[:]?", re.IGNORECASE)
    questions, errors = [], []

    for idx, block in enumerate(blocks, 1):
        lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
        m_q = re.match(r"^\d+\.\s*(.+)", lines[0])
        q_text = m_q.group(1).strip() if m_q else lines[0]

        corr_idx = next((i for i, ln in enumerate(lines) if corr_pattern.match(ln)), None)
        if corr_idx is None:
            errors.append((idx, "Не знайдено рядок із правильними відповідями"))
            continue

        ans_lines = lines[1:corr_idx]
        corr_line = lines[corr_idx]
        letters = re.findall(r"[A-ГA-D]", corr_line)
        correct_set = {mapping.get(l, l.upper()) for l in letters}

        answers = []
        if len(ans_lines) == 1 and re.search(r"(?i)true", ans_lines[0]):
            is_true = re.search(r"(?i)true", corr_line) is not None
            answers = [("true", is_true), ("false", not is_true)]
        else:
            for ln in ans_lines:
                m_a = re.match(r"^([A-Г])[\)\.]*\s*(.+)", ln)
                if not m_a:
                    errors.append((idx, f"Невірний формат відповіді: '{ln}'"))
                    break
                letter = mapping.get(m_a.group(1), m_a.group(1).upper())
                answers.append((m_a.group(2).strip(), letter in correct_set))
            if len(answers) < 2:
                errors.append((idx, "Менше двох варіантів відповіді"))
                continue

        questions.append({"text": q_text, "answers": answers})

    return questions, errors

def parse_from_excel(uploaded_file):
    """Парсер тестів з Excel."""
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    items = []
    for cell in ws['A']:
        txt = str(cell.value).strip() if cell.value else ""
        is_corr = False
        if cell.fill and getattr(cell.fill, 'fill_type', None) == 'solid':
            if getattr(cell.fillstart_color, 'rgb', '').endswith('FFFF00'):
                is_corr = True
        items.append((txt, is_corr))

    blocks, curr = [], []
    for txt, corr in items:
        if not txt and curr:
            blocks.append(curr)
            curr = []
        elif txt:
            curr.append((txt, corr))
    if curr:
        blocks.append(curr)

    questions, errors = [], []
    for idx, blk in enumerate(blocks, 1):
        if len(blk) < 3:
            errors.append((idx, "Потрібно питання + ≥2 відповіді"))
            continue
        questions.append({"text": blk[0][0], "answers": blk[1:]})
    return questions, errors

def parse_from_word(uploaded_file):
    """Парсер тестів з Word (.docx)."""
    doc = Document(uploaded_file)
    answer_pattern = re.compile(r'^[A-ЯA-Z]\.\s*', re.U)
    questions, errors = [], []
    curr_q = None

    for para in doc.paragraphs:
        for line in para.text.splitlines():
            txt = line.strip()
            if not txt:
                continue

            # Inline формат: "Питання: A; B; C;"
            if ':' in txt and txt.count(';') >= 2 and not answer_pattern.match(txt):
                if curr_q:
                    questions.append(curr_q)
                part_q, part_ans = txt.split(':', 1)
                curr_q = {"text": part_q.strip(), "answers": []}
                segments = [seg.strip().rstrip(';') for seg in part_ans.split(';') if seg.strip()]
                for seg in segments:
                    is_corr = any(run.bold and seg in run.text for run in para.runs)
                    curr_q["answers"].append((seg, is_corr))

            # Окремі абзаци-відповіді "A. Відповідь"
            elif answer_pattern.match(txt):
                if curr_q is None:
                    errors.append(f"Відповідь без питання: «{txt}»")
                    continue
                is_corr = any(run.bold for run in para.runs)
                ans_txt = answer_pattern.sub('', txt)
                curr_q["answers"].append((ans_txt, is_corr))

            else:
                if curr_q is None:
                    curr_q = {"text": txt, "answers": []}
                elif curr_q["answers"]:
                    questions.append(curr_q)
                    curr_q = {"text": txt, "answers": []}
                else:
                    curr_q["text"] += " " + txt

    if curr_q:
        questions.append(curr_q)

    return questions, errors

# ================== Інтерфейс режимів ==================

mode = st.sidebar.selectbox("Виберіть режим створення тесту", [
    "1. Excel",
    "2. По тексту (GPT)",
    "3. Вручну",
    "4. Готовий тест",
    "5. Word → XML",
    "6. YouTube → XML"
])

# 1️⃣ Excel
if mode == "1. Excel":
    st.header("1️⃣ Режим Excel")
    uploaded = st.file_uploader("Завантажте .xlsx", type=["xlsx"])
    if uploaded:
        qs, errs = parse_from_excel(uploaded)
        if errs:
            st.error("Помилки при парсингу Excel:")
            for i, m in errs:
                st.write(f"- Блок {i}: {m}")
        elif not qs:
            st.warning("Питань не знайдено.")
        else:
            st.success(f"Знайдено {len(qs)} питань")
            if st.button("Генерувати XML"):
                xml_str = generate_moodle_xml_string(qs)
                st.subheader("📄 Попередній перегляд XML")
                st.text_area("", xml_str, height=300)
                download_xml(xml_str, "excel_test.xml")

# 2️⃣ GPT
elif mode == "2. По тексту (GPT)":
    st.header("2️⃣ Режим GPT-генерації")
    user_text = st.text_area("Вставте текст для генерації тесту українською", height=200)
    if st.button("Створити тест"):
        prog = st.progress(0)
        status = st.empty()
        status.text("1/3: Підготовка запиту…"); prog.progress(10)
        system_prompt = (
            """Ви — асистент із жорстким обмеженням на 10 питань у форматі Moodle XML українською мовою.
1) Використайте тільки наданий текст.
2) Створіть **саме 10** логічних питань українською:
   – 4–5 питань Single-choice,
   – 2–3 питання True/False,
   – 2–3 питання Multiple-choice.
3) Кожне питання (окрім True/False) має мати 4 варіанти (A, B, C, D).
4) Перед поверненням перевірте, що загальна кількість питань = 10.
5) Поверніть **тільки** XML-код (без коментарів) і одразу припиніть після 10-го питання.
<END>
"""
        )
        status.text("2/3: Відправка запиту…"); prog.progress(40)
        resp = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_text}
            ],
            temperature=0
        )
        status.text("3/3: Обробка відповіді…"); prog.progress(70)
        text = resp.choices[0].message.content.strip()
        # Видаляємо можливі markdown-обгортки ```xml … ```
        text = re.sub(r"^```(?:xml)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)

        if text.startswith("<?xml") or text.lstrip().startswith("<quiz"):
            count = len(re.findall(r'<question', text))
            if count != 10:
                st.error(f"GPT повернув {count} питань замість 10.")
            xml_str = text if text.startswith("<?xml") else '<?xml version="1.0" encoding="UTF-8"?>\n' + text
            if count > 10:
                parts = re.findall(r'(<question[\s\S]*?</question>)', xml_str)
                xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n<quiz>' + ''.join(parts[:10]) + '</quiz>'
                st.warning("Обрізано до перших 10 питань.")
            elif count < 10:
                st.error(f"GPT згенерував тільки {count} питань, а потрібно 10.")
            st.subheader("📄 Попередній перегляд XML")
            st.code(xml_str)
            download_xml(xml_str, "gpt_test.xml")
            status.text("Готово!"); prog.progress(100)
        else:
            status.text("3/3: Парсинг відповіді…"); prog.progress(80)
            qs, errs = parse_text_format(text)
            if errs:
                st.error("Помилки при парсингу GPT-відповіді:")
                for i, m in errs:
                    st.write(f"- Блок {i}: {m}")
            else:
                if len(qs) != 10:
                    st.error(f"GPT згенерував {len(qs)} питань замість 10.")
                    qs = qs[:10]
                status.text("4/4: Генерація XML…"); prog.progress(100)
                xml_str = generate_moodle_xml_string(qs)
                st.subheader("📄 Попередній перегляд XML")
                st.code(xml_str)
                download_xml(xml_str, "gpt_test.xml")

# 3️⃣ Вручну
elif mode == "3. Вручну":
    st.header("3️⃣ Ручне створення питань")
    if "manual_qs" not in st.session_state:
        st.session_state.manual_qs = []
    qtype = st.selectbox("Тип питання", ["Single-choice", "Multiple-choice", "True/False"])
    with st.form("manual_form"):
        q_txt = st.text_input("Текст питання")
        answers = []
        if qtype == "True/False":
            corr = st.radio("Правильна відповідь", ["true", "false"])
            answers = [("true", corr == "true"), ("false", corr == "false")]
        else:
            cols = st.columns([8, 1])
            with cols[0]:
                a1 = st.text_input("A."); a2 = st.text_input("B.")
                a3 = st.text_input("C."); a4 = st.text_input("D.")
            with cols[1]:
                c1 = st.checkbox("", key="c1"); c2 = st.checkbox("", key="c2")
                c3 = st.checkbox("", key="c3"); c4 = st.checkbox("", key="c4")
            answers = [(a1, c1), (a2, c2), (a3, c3), (a4, c4)]
        submitted = st.form_submit_button("Додати питання")
        if submitted:
            st.session_state.manual_qs.append({"text": q_txt, "answers": answers})
    if st.session_state.manual_qs:
        st.subheader("Список питань")
        for idx, q in enumerate(st.session_state.manual_qs, 1):
            st.write(f"{idx}. {q['text']}")
        if st.button("Генерувати XML для ручних питань"):
            xml_str = generate_moodle_xml_string(st.session_state.manual_qs)
            st.text_area("", xml_str, height=300)
            download_xml(xml_str, "manual_test.xml")

# 4️⃣ Готовий тест
elif mode == "4. Готовий тест":
    st.header("4️⃣ Режим готового тесту")
    ready = st.text_area(
        "Вставте готовий тест у зазначеному форматі (1. … A. … Правильний відповідь: …)",
        height=200
    )
    if st.button("Генерувати XML з готового тесту", key="gen_ready"):
        qs, errs = parse_text_format(ready)
        if errs:
            st.error("Помилки при розборі тесту:")
            for idx, m in errs:
                st.write(f"- Блок {idx}: {m}")
        else:
            st.success(f"Знайдено {len(qs)} питань — генеруємо XML…")
            xml_str = generate_moodle_xml_string(qs)
            st.text_area("", xml_str, height=300, label_visibility="collapsed")
            download_xml(xml_str, "ready_test.xml")

# 5️⃣ Word → XML
elif mode == "5. Word → XML":
    st.header("5️⃣ Режим Word → Moodle XML")
    file = st.file_uploader("Завантажте .docx з жирними правильними відповідіми", type=["docx"])
    if file:
        qs, errs = parse_from_word(file)
        if errs:
            st.error("Проблеми з розбором Word-документу:")
            for e in errs:
                st.write(f"- {e}")
        elif not qs:
            st.warning("Не знайдено жодного питання в документі.")
        else:
            st.success(f"Розпізнано {len(qs)} питань.")
            if st.button("Генерувати Moodle XML"):
                xml_str = generate_moodle_xml_string(qs)
                st.subheader("📄 Попередній перегляд XML")
                st.text_area("", xml_str, height=300)
                download_xml(xml_str, "word_test.xml")

# 6️⃣ YouTube → XML
elif mode == "6. YouTube → XML":
    st.header("6️⃣ Режим YouTube → Moodle XML")
    yt_url = st.text_input("Вставте посилання на YouTube відео")
    if st.button("Створити тест з відео"):
        prog = st.progress(0)
        status = st.empty()
        try:
            status.text("1/4: Завантаження аудіо з YouTube…"); prog.progress(10)
            audio_path = download_audio_from_youtube(yt_url)

            status.text("2/4: Транскрибація аудіо…"); prog.progress(30)
            with open(audio_path, 'rb') as audio_file:
                transcript_resp = openai.Audio.transcribe("whisper-1", audio_file)
            transcript_text = transcript_resp["text"]

            status.text("3/4: Генерація питань GPT…"); prog.progress(60)
            system_prompt = (
                """Ви — асистент із жорстким обмеженням на 10 питань у форматі Moodle XML українською мовою.
1) Використайте тільки наданий текст.
2) Створіть **саме 10** логічних питань українською:
   – 4–5 питань Single-choice,
   – 2–3 питання True/False,
   – 2–3 питання Multiple-choice.
3) Кожне питання (окрім True/False) має мати 4 варіанти (A, B, C, D).
4) Перед поверненням перевірте, що загальна кількість питань = 10.
5) Поверніть **тільки** XML-код (без коментарів) і одразу припиніть після 10-го питання.
<END>
"""
            )
            resp = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": transcript_text}
                ],
                temperature=0
            )
            status.text("4/4: Обробка відповіді…"); prog.progress(80)
            xml_content = resp.choices[0].message.content.strip()
            # Видаляємо можливі markdown-обгортки ```xml … ```
            xml_content = re.sub(r"^```(?:xml)?\s*", "", xml_content)
            xml_content = re.sub(r"\s*```$", "", xml_content)

            if not xml_content.startswith("<?xml"):
                xml_content = "<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n" + xml_content

            st.subheader("📄 Попередній перегляд XML")
            st.code(xml_content)
            download_xml(xml_content, "youtube_test.xml")
            status.text("Готово!"); prog.progress(100)

        except Exception as e:
            st.error(f"Помилка: {e}")
